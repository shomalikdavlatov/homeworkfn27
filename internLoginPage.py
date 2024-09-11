from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QWidget, QMainWindow, QPushButton, QLabel, QLineEdit, QVBoxLayout, QHBoxLayout,QFormLayout, QMessageBox
import pickle
from openpyxl import Workbook


class Login(QMainWindow):
    def __init__(self, parent):
        super().__init__(parent)
        self.setGeometry(340, 200, 680, 400)
        self.setWindowTitle("Login")

        central = QWidget()
        self.setCentralWidget(central)
        mainLayout = QVBoxLayout()

        self.btnBack = QPushButton("Back")
        self.btnBack.clicked.connect(lambda: self.btnBackClicked())

        headerLayout = QHBoxLayout()
        headerLayout.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        headerLayout.addWidget(self.btnBack)

        username = QLabel("Username:")
        self.username = QLineEdit()
        miniLayout1 = QFormLayout()
        miniLayout1.addRow(username, self.username)

        password = QLabel("Password:")
        self.password = QLineEdit()
        miniLayout2 = QFormLayout()
        miniLayout2.addRow(password, self.password)

        self.btnSignIn = QPushButton("Sign In")
        self.btnSignIn.clicked.connect(lambda: self.btnSignInClicked())
        
        mainLayout.addLayout(headerLayout)
        mainLayout.addLayout(miniLayout1)
        mainLayout.addLayout(miniLayout2)
        mainLayout.addWidget(self.btnSignIn)

        central.setLayout(mainLayout)

        self.show()

    def btnBackClicked(self):
        self.close()
        self.parent().show()

    def btnSignInClicked(self):
        username = self.username.text()
        password = self.password.text()
        try:
            with open('users.bin', 'rb') as file:
                users = pickle.load(file)
                noUser = True
                for user in users:
                    if user[0] == username and user[3] == password:
                        self.signedIn()
                        noUser = False
                if noUser:
                    self.noUser()
        except:
            self.noUser()

    def noUser(self):
        info = QMessageBox()
        info.critical(self, 'critical', 'The user is not found!', QMessageBox.StandardButton.Ok)

    def signedIn(self):
        info = QMessageBox()
        info.setIcon(QMessageBox.Icon.Question)
        info.setText("Foydalanuvchilar ma'lumotlarini exel filega yuklashni xohlaysizmi?")
        info.setStandardButtons(QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)

        info.buttonClicked.connect(self.btnClicked)

        info.exec_()

    def btnClicked(self, btn):
        if btn.text() == 'OK':
            workbook = Workbook()
            sheet = workbook.active

            sheet.cell(1, 1).value = 'username'
            sheet.cell(1, 2).value = 'full name'
            sheet.cell(1, 3).value = 'phone number'
            sheet.cell(1, 4).value = 'password'

            with open('users.bin', 'rb') as file:
                users = pickle.load(file)

                for i in range(len(users)):
                    sheet.cell(i+2, 1).value = users[i][0]
                    sheet.cell(i+2, 2).value = users[i][1]
                    sheet.cell(i+2, 3).value = users[i][2]
                    sheet.cell(i+2, 4).value = users[i][3]

            workbook.save("users.xlsx")

